import sys
import traceback
import datetime
import requests
import openpyxl
from varible import *


def userInput(msg):
    print(msg)
    name = input("請輸入你的名字 : ")
    date = input("請輸入欲驗證的日期 : ")
    while True:
        try:
            if date != datetime.datetime.strptime(date, "%m-%d").strftime("%m-%d"):
                raise ValueError
            else:
                return name, date
        except ValueError:
            date = input("你輸入的日期格式有誤OHO\n請重新輸入日期 : ")


def getSunTime(uD):
    lastDate = (datetime.datetime.strptime(uD, "%Y-%m-%d")-datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    nextDate = (datetime.datetime.strptime(uD, "%Y-%m-%d")+datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    url = BaseURL + sun_data_url
    payload = {
        'Authorization': "CWB-5FBB0B6F-E22D-423A-87B9-2719895FDAF9",
        'limit': 1,
        'format': "JSON",
        'locationName': ["花蓮縣"],
        'dataTime': [lastDate, uD, nextDate],
    }
    data = requests.get(url, params=payload)
    data_json = data.json()
    location = data_json['records']['locations']["location"][0] # 花蓮縣
    lastNight = location['time'][0]['parameter'][7]['parameterValue'] # 上一日 民用暮光終
    lastNight = lastDate[5:] + " " + lastNight
    todayDay = location['time'][1]['parameter'][0]['parameterValue'] # 當日 民用曙光始
    todayDay = uD[5:] + " " + todayDay
    todayNight = location['time'][1]['parameter'][7]['parameterValue'] # 當日 民用暮光終
    todayNight = uD[5:] + " " + todayNight
    nextDay = location['time'][2]['parameter'][0]['parameterValue'] # 下一日 民用曙光始
    nextDay = nextDate[5:] + " " + nextDay
    return lastNight, todayDay, todayNight, nextDay


def catchError(e):
    error_class = e.__class__.__name__
    detail = e.args[0]
    cl, exc, tb = sys.exc_info()
    lastCallStack = traceback.extract_tb(tb)
    fileName = lastCallStack[0]
    lineNum = lastCallStack[1]
    funcName = lastCallStack[2]
    errMsg = "File \"{}\", \nline {}, \nin {}: [{}] {}\n".format(fileName, lineNum, funcName, error_class, detail)
    print(errMsg)
    sys.exit(1)


def openExcelFile(op):
    wb = openpyxl.load_workbook(op, data_only=False)
    wb.active = 0
    ws = wb.active
    return ws, wb


def initExcelFile(uN, uD, mode):
    # todo now YEAR=2022
    uD = '2022-'+uD
    tl = getSunTime(uD)
    input_file = ""
    output_file = ""
    if mode == "AG":
        input_file = inputFileAG
        output_file = outputFileAG
    if mode == "PO":
        input_file = inputFilePO
        output_file = outputFilePO
    if mode == "SL":
        input_file = inputFileSL
        output_file = outputFileSL
    ws, wb = openExcelFile(input_file)
    ws['A1'].value = '驗測日期：{}'.format(datetime.date.today())
    ws['A2'].value = '影片日期：{}'.format(uD)
    ws['A3'].value = '驗測人員：{}'.format(uN)
    ws3 = wb["時段與氣候"]
    ws3['B3'].value = ("{} ~ {}".format(tl[0], tl[1])).replace("-", "/")
    ws3['B4'].value = ("{} ~ {}".format(tl[1], tl[2])).replace("-", "/")
    ws3['B5'].value = ("{} ~ {}".format(tl[2], tl[3])).replace("-", "/")
    ws3['B6'].value = ("2022-{}".format(tl[1])).replace("-", "/")
    wb.save(output_file)



